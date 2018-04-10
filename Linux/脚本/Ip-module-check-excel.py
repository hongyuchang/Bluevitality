#!/usr/bin/env python
#coding=utf-8

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
import sys

HOSTS=list()
MODULE=list()

#------------------------------- 读取
rb = load_workbook(filename=u'in.xlsx')
print rb.sheetnames
rs=rb.get_sheet_by_name(rb.sheetnames[2]) 

#获取所有行
rows = rs.rows

#迭代所有行的指定列
for row in rows:
	line = [col.value for col in row]
	HOSTS.append(line[1])
	MODULE.append(line[3])	

#------------------------------- 写入
wb = Workbook()
ws = wb.active

#每个模块的LOG检查关键字
CHECK_STR=[
"OnestBusiFileUploadFailed:fileType=P",
"OnestBusiFileDownloadFailed:fileType=P",
"OnestBusiFileUploadFailed:fileType=W",
"OnestBusiFileDownloadFailed:fileType=W",
"OnestBusiFileUploadFailed:fileType=V",
"OnestBusiFileDownloadFailed:fileType=V",
"OnestGztFileUploadFailed:fileType=GZT",
"OnestGztFileDownloadFailed:fileType=GZT",
"RnfsBusiFileUploadFailed",
"RnfsBusiFileDownloadFailed",
"RnfsGztFileUploadFailed:fileType=BusiFile",
"RnfsGztFileDownloadFailed:fileType=BusiFile",
"sendMqMsghasFaild",
"reqLinkfacefailedCode",
"initJvmCacheDataFailed",
"initRedisCacheDataFailed",
"Send message to vertica TOPIC error",
"BusinessFlowController GeneralException",
"BusinessFlowController Exception",
"BusinessFlowController error"
]

x=1
for a in xrange(len(HOSTS)):
	for b in xrange(len(CHECK_STR)):
		ws.cell(row=int(x),column=1).value=HOSTS[a]	#根据关键字个数输出N个相同的主机
		ws.cell(row=int(x),column=2).value=MODULE[a]	#根据关键字个数输出N个相同的模块地址
		ws.cell(row=int(x),column=3).value=CHECK_STR[b]	#输出关键字
		x+=1
if __name__ == '__main__':
        wb.save(filename="1.xlsx")
